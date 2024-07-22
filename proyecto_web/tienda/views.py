from django.shortcuts import render
from tienda.models import Producto
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from django.http.response import HttpResponse
from tienda.utils import render_pdf

def tienda(request):
    product = Producto.objects.all()
    return render(request, 'tienda.html', {'productos': product})

def reporte(request):
    product = Producto.objects.all()
    return render(request, 'Listado.html', {'productos': product})

class ReporteProductoExcel(TemplateView):
    def get(self, request, *args, **Kwargs):
        producto = Producto.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de productos'
        ws['A1'] = 'Reporte productos'
        ws.merge_cells('A1:C1')

        ws['A3'] = 'Producto'
        ws['B3'] = 'Precio'
        ws['C3'] = 'Imagen'
        cont = 4

        for product in producto:
            ws.cell(row=cont, column=1).value = product.nombre
            ws.cell(row=cont, column=2).value = product.precio
            ws.cell(row=cont, column=3).value = product.imagen.url
            cont += 1
        nombre_archivo = 'ReporteProductosExcel.xlsx'
        response = HttpResponse(content_type = "aplication/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['content-Disposition'] = content
        wb.save(response)
        return response

def Listado(request):
    product = Producto.objects.all()
    return render(request, 'Listado.html', {'productos': product})

class ListaProductoPdf(View):
    def get(self, request, *args, **Kwargs):
        productos = Producto.objects.all()
        data = {
            "productos": productos
        }
        pdf = render_pdf('tienda/Listado.html', data)
        return HttpResponse(pdf, content_type = 'application/pdf')